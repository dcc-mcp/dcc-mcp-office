"""office-generate-production-dashboard / generate_dashboard — Workbook IR → XLSX.

Parameter resolution order (dcc-mcp-core execute_script convention):
1. stdin JSON: {"input": ..., "output_dir": ...}
2. CLI flags: --input --out
"""

from __future__ import annotations

import argparse
import json
import re
import sys
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.drawing.image import Image as WorksheetImage
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

IR_VERSION = "office-ir/1.0"
HEADER_FILL = PatternFill("solid", fgColor="0B1020")
HEADER_FONT = Font(color="FFFFFF", bold=True)
ALT_FILL = PatternFill("solid", fgColor="F4F2ED")
ACCENT_FILL = PatternFill("solid", fgColor="3157F6")
CARD_FILL = PatternFill("solid", fgColor="F4F2ED")
CARD_ALT_FILL = PatternFill("solid", fgColor="E9F2FF")
INK = "172033"
MUTED = "5E6678"
ACCENT = "3157F6"
CORAL = "FF6B5E"
WHITE = "FFFFFF"
GRID_BORDER = Border(
    left=Side(style="thin", color="D0D5DD"),
    right=Side(style="thin", color="D0D5DD"),
    top=Side(style="thin", color="D0D5DD"),
    bottom=Side(style="thin", color="D0D5DD"),
)


class WorkbookIrError(ValueError):
    """Workbook IR contract violation with a json-path hint."""


def _require(mapping: dict, key: str, path: str) -> Any:
    if key not in mapping:
        raise WorkbookIrError(f"[{path}] missing required key '{key}'")
    return mapping[key]


def load_workbook_ir(path: str | Path) -> dict:
    raw_path = Path(path)
    if not raw_path.is_file():
        raise WorkbookIrError(f"[$] input file not found: {raw_path}")
    raw = json.loads(raw_path.read_text(encoding="utf-8"))
    if raw.get("schema_version") != IR_VERSION:
        raise WorkbookIrError(f"[$] expected schema_version '{IR_VERSION}'")
    if raw.get("kind") != "workbook":
        raise WorkbookIrError("[$.kind] expected 'workbook'")
    document = _require(raw, "document", "$")
    worksheets = document.get("worksheets", [])
    if not worksheets:
        raise WorkbookIrError("[$.document.worksheets] must be a non-empty list")
    for wi, ws in enumerate(worksheets):
        path = f"$.document.worksheets[{wi}]"
        if not ws.get("name"):
            raise WorkbookIrError(f"[{path}] missing worksheet 'name'")
        for ti, table in enumerate(ws.get("tables", [])):
            tpath = f"{path}.tables[{ti}]"
            headers = table.get("headers")
            rows = table.get("rows")
            if not isinstance(headers, list) or not headers:
                raise WorkbookIrError(f"[{tpath}] 'headers' must be a non-empty list")
            if not isinstance(rows, list) or not all(isinstance(r, list) for r in rows):
                raise WorkbookIrError(f"[{tpath}] 'rows' must be a list of lists")
    return raw


class DashboardBuilder:
    """Builds one XLSX from a Workbook IR. SRP: layout, formulas, chart."""

    def __init__(self, ir: dict, *, visual_asset: Path | None = None) -> None:
        self.ir = ir
        self.visual_asset = visual_asset
        self.wb = Workbook()
        self.wb.remove(self.wb.active)
        self.summary: dict[str, Any] = {"sheets": [], "rows": 0}
        self._table_starts: dict[str, int] = {}
        self._table_ends: dict[str, int] = {}
        self._chart_sheets: set[str] = set()

    def build(self) -> Workbook:
        metadata = self.ir.get("metadata", {})
        self.wb.properties.title = metadata.get("title", "Production dashboard")
        self.wb.properties.creator = metadata.get("author", "DCC-MCP")
        self.wb.properties.subject = "Editable Office automation capability dashboard"
        self.wb.calculation.fullCalcOnLoad = True
        self.wb.calculation.forceFullCalc = True
        for ws_spec in self.ir["document"]["worksheets"]:
            self._build_sheet(ws_spec)
        for chart_spec in self.ir["document"].get("charts", []):
            self._build_chart(chart_spec)
        for worksheet in self.wb.worksheets:
            self._configure_print(worksheet)
        return self.wb

    def _build_sheet(self, spec: dict) -> None:
        ws = self.wb.create_sheet(title=spec["name"])
        ws.sheet_view.showGridLines = False
        ws.sheet_properties.tabColor = ACCENT
        self._write_title_band(ws, spec["name"])
        self._add_visual_asset(ws)
        sheet_rows = 0
        cursor = 9
        for index, table in enumerate(spec.get("tables", [])):
            start, end = self._write_table(ws, table, cursor)
            if index == 0:
                self._table_starts[ws.title] = start
                self._table_ends[ws.title] = end
                self._write_kpi_cards(ws, table, start, end)
            cursor = end + 3
            sheet_rows += len(table["rows"])
        if ws.title == "Workflow Packs":
            self._write_workflow_rail(ws)
        ws.freeze_panes = "A10"
        self.summary["sheets"].append({"name": spec["name"], "tables": len(spec.get("tables", [])), "rows": sheet_rows})
        self.summary["rows"] += sheet_rows

    def _write_title_band(self, ws, sheet_name: str) -> None:
        title = self.ir.get("metadata", {}).get("title", "Production dashboard")
        ws.merge_cells("A1:M1")
        ws["A1"] = title
        ws["A1"].fill = HEADER_FILL
        ws["A1"].font = Font(name="Aptos Display", size=25, bold=True, color=WHITE)
        ws["A1"].alignment = Alignment(vertical="center")
        ws.row_dimensions[1].height = 43

        ws.merge_cells("A2:M2")
        ws["A2"] = f"{sheet_name.upper()}  /  WORKBOOK IR  /  NATIVE EXCEL PROOF"
        ws["A2"].fill = ACCENT_FILL
        ws["A2"].font = Font(name="Aptos", size=10, color=WHITE)
        ws["A2"].alignment = Alignment(vertical="center")
        ws.row_dimensions[2].height = 24
        ws.row_dimensions[3].height = 8

    def _add_visual_asset(self, ws) -> None:
        if ws.title != "Runtime Surface" or self.visual_asset is None:
            return
        if not self.visual_asset.is_file():
            raise WorkbookIrError(f"[$.metadata.visual_asset] file not found: {self.visual_asset}")
        visual = WorksheetImage(str(self.visual_asset))
        for column in range(8, 14):
            ws.column_dimensions[get_column_letter(column)].width = 10.5
        visual.width = 405
        visual.height = 148
        visual.anchor = "H4"
        ws.add_image(visual)
        ws.merge_cells("H9:M9")
        note = ws["H9"]
        note.value = "ORIGINAL VISUAL  ·  MEASURED CAPABILITY LANDSCAPE"
        note.font = Font(name="Aptos", size=7.5, bold=True, color=MUTED)
        note.alignment = Alignment(horizontal="right", vertical="center")

    def _write_kpi_cards(self, ws, table: dict, start: int, end: int) -> None:
        headers = table["headers"]
        rows = table["rows"]
        cards: list[tuple[str, Any]] = []
        if "Live capabilities" in headers:
            app_index = headers.index("Application")
            value_index = headers.index("Live capabilities")
            cards.extend((str(row[app_index]), row[value_index]) for row in rows[:3])
            value_column = get_column_letter(value_index + 1)
            cards.append(
                (
                    "Total live",
                    f'=SUM({value_column}{start + 1}:{value_column}{end})',
                )
            )
        else:
            cards = [
                ("Workflow packs", len(rows)),
                ("Editable outputs", len(rows)),
                ("Governed contract", 1),
                ("Verified rows", f"=COUNTA(A{start + 1}:A{end})"),
            ]

        visual_layout = ws.title == "Runtime Surface" and self.visual_asset is not None
        for index, (label, value) in enumerate(cards[:4]):
            if visual_layout:
                first_col = 1 + index * 2
                last_col = first_col + 1 if index < 3 else first_col
            else:
                first_col = 1 + index * 3
                last_col = first_col + 2
            label_range = f"{get_column_letter(first_col)}4:{get_column_letter(last_col)}4"
            value_range = f"{get_column_letter(first_col)}5:{get_column_letter(last_col)}8"
            ws.merge_cells(label_range)
            ws.merge_cells(value_range)
            label_cell = ws.cell(row=4, column=first_col, value=str(label).upper())
            label_cell.fill = PatternFill("solid", fgColor=CORAL) if index == 3 else CARD_FILL
            label_cell.font = Font(name="Aptos", size=8, bold=True, color=WHITE if index == 3 else MUTED)
            label_cell.alignment = Alignment(horizontal="center", vertical="center")
            value_cell = ws.cell(row=5, column=first_col, value=value)
            value_cell.fill = PatternFill("solid", fgColor=CORAL) if index == 3 else CARD_ALT_FILL
            value_cell.font = Font(name="Aptos Display", size=26, bold=True, color=WHITE if index == 3 else ACCENT)
            value_cell.alignment = Alignment(horizontal="center", vertical="center")
            label_cell.border = GRID_BORDER
            value_cell.border = GRID_BORDER
        ws.row_dimensions[4].height = 22
        ws.row_dimensions[5].height = 22
        ws.row_dimensions[6].height = 22
        ws.row_dimensions[7].height = 22
        ws.row_dimensions[8].height = 22

    def _write_table(self, ws, table: dict, start: int) -> tuple[int, int]:
        headers = table["headers"]
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=start, column=col, value=header)
            cell.fill = HEADER_FILL
            cell.font = Font(name="Aptos", size=10, color=WHITE, bold=True)
            cell.alignment = Alignment(horizontal="center")
            cell.border = GRID_BORDER
        for row_idx, row in enumerate(table["rows"], start=start + 1):
            for col_idx, value in enumerate(row, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                if row_idx % 2 == 0:
                    cell.fill = ALT_FILL
                cell.font = Font(name="Aptos", size=10, color=INK)
                cell.alignment = Alignment(vertical="top", wrap_text=True)
                cell.border = GRID_BORDER
            ws.row_dimensions[row_idx].height = 32
        ws.row_dimensions[start].height = 25
        end = start + len(table["rows"])

        display_name = _safe_table_name(f"{ws.title}_{table.get('name', 'table')}")
        excel_table = Table(displayName=display_name, ref=f"A{start}:{get_column_letter(len(headers))}{end}")
        excel_table.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium2",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        ws.add_table(excel_table)

        for col, header in enumerate(headers, start=1):
            values = [str(header), *(str(row[col - 1]) for row in table["rows"])]
            width = min(max(max(len(value) for value in values) + 3, 14), 42)
            ws.column_dimensions[get_column_letter(col)].width = width
        return start, end

    def _write_workflow_rail(self, ws) -> None:
        steps = [
            ("F10:G12", "01\nSOURCE\nReusable IR", ACCENT),
            ("I10:J12", "02\nGOVERN\nPolicy + checkpoint", "0D9488"),
            ("L10:M12", "03\nPROVE\nNative export", CORAL),
        ]
        for cell_range, label, fill in steps:
            ws.merge_cells(cell_range)
            cell = ws[cell_range.split(":", maxsplit=1)[0]]
            cell.value = label
            cell.fill = PatternFill("solid", fgColor=fill)
            cell.font = Font(name="Aptos Display", size=12, bold=True, color=WHITE)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        for address in ("H11", "K11"):
            cell = ws[address]
            cell.value = "→"
            cell.font = Font(name="Aptos Display", size=22, bold=True, color=ACCENT)
            cell.alignment = Alignment(horizontal="center", vertical="center")
        for column in range(6, 14):
            ws.column_dimensions[get_column_letter(column)].width = 10.5

    def _build_chart(self, spec: dict) -> None:
        if spec.get("type") != "bar":
            raise WorkbookIrError(f"[$.document.charts] unsupported chart type '{spec.get('type')}'")
        if spec["sheet"] not in self.wb.sheetnames:
            raise WorkbookIrError(f"[$.document.charts] unknown sheet '{spec['sheet']}'")
        ws = self.wb[spec["sheet"]]
        if ws.title not in self._table_starts:
            raise WorkbookIrError(f"[$.document.charts] sheet '{ws.title}' has no source table")
        table_start = self._table_starts[ws.title]
        value_start, value_end = _ref_rows(spec["values"])
        category_start, category_end = _ref_rows(spec["categories"])
        chart = BarChart()
        chart.title = spec.get("title", "Chart")
        chart.style = 13
        chart.type = "bar"
        chart.height = 7.1
        chart.width = 14.8
        chart.legend = None
        chart.x_axis.delete = True
        chart.y_axis.delete = True
        chart.y_axis.majorGridlines = None
        chart.varyColors = False
        chart.gapWidth = 45
        chart.dLbls = DataLabelList()
        chart.dLbls.showVal = True
        chart.dLbls.showCatName = True
        chart.dLbls.showSerName = False
        chart.dLbls.showLegendKey = False
        chart.dLbls.showPercent = False
        chart.dLbls.dLblPos = "outEnd"
        data = Reference(
            ws,
            min_col=_ref_col(spec["values"]),
            min_row=table_start + value_start - 1,
            max_row=table_start + value_end - 1,
        )
        cats = Reference(
            ws,
            min_col=_ref_col(spec["categories"]),
            min_row=table_start + category_start - 1,
            max_row=table_start + category_end - 1,
        )
        chart.add_data(data, titles_from_data=False)
        chart.set_categories(cats)
        series = chart.series[0]
        series.graphicalProperties.solidFill = ACCENT
        series.graphicalProperties.line.noFill = True
        ws.add_chart(chart, spec.get("anchor", "F9"))
        self._chart_sheets.add(ws.title)

    def _configure_print(self, ws) -> None:
        ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
        ws.page_setup.paperSize = ws.PAPERSIZE_LETTER
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 1
        ws.sheet_properties.pageSetUpPr.fitToPage = True
        ws.page_margins.left = 0.32
        ws.page_margins.right = 0.32
        ws.page_margins.top = 0.42
        ws.page_margins.bottom = 0.42
        ws.page_margins.header = 0.2
        ws.page_margins.footer = 0.2
        ws.oddFooter.left.text = "dcc-mcp-office showcase"
        ws.oddFooter.center.text = "Page &P of &N"
        ws.oddFooter.right.text = "Generated from Workbook IR"
        last_row = 24 if ws.title in self._chart_sheets else max(self._table_ends.get(ws.title, 12) + 3, 15)
        ws.print_area = f"A1:M{last_row}"


def _ref_col(ref: str) -> int:
    """A1-style column letters → index (A=1)."""
    start, _, _ = ref.partition(":")
    letters = "".join(ch for ch in start if ch.isalpha())
    value = 0
    for ch in letters.upper():
        value = value * 26 + (ord(ch) - ord("A") + 1)
    return value


def _ref_row(ref: str) -> int:
    """A1-style row digits → int."""
    return int("".join(ch for ch in ref if ch.isdigit()))


def _ref_rows(ref: str) -> tuple[int, int]:
    start, _, end = ref.partition(":")
    start_row = _ref_row(start)
    return start_row, _ref_row(end) if end else start_row


def _safe_table_name(value: str) -> str:
    cleaned = re.sub(r"[^A-Za-z0-9_]", "_", value)
    if not cleaned or cleaned[0].isdigit():
        cleaned = f"Table_{cleaned}"
    return cleaned[:255]


def artifact_stem(document_id: str) -> str:
    """Safe filesystem stem for a document id (ids may contain ':')."""
    return re.sub(r"[^A-Za-z0-9._-]+", "-", document_id).strip("-") or "dashboard"


def run(params: dict) -> None:
    ir = load_workbook_ir(params["input"])
    input_path = Path(params["input"])
    visual_asset = None
    configured_asset = ir.get("metadata", {}).get("visual_asset")
    if configured_asset:
        visual_asset = input_path.parent / str(configured_asset)
    builder = DashboardBuilder(ir, visual_asset=visual_asset)
    workbook = builder.build()
    out_dir = Path(params.get("output_dir", "output"))
    out_dir.mkdir(parents=True, exist_ok=True)
    out_path = out_dir / f"{artifact_stem(ir.get('document_id', 'dashboard'))}.xlsx"
    workbook.save(str(out_path))
    ok = out_path.is_file() and out_path.stat().st_size > 0
    print(
        json.dumps(
            {
                "success": ok,
                "message": f"dashboard '{ir.get('document_id', 'dashboard')}' generated",
                "context": {"artifact": str(out_path), "summary": builder.summary},
            },
            ensure_ascii=False,
        )
    )


def _force_utf8_stdio() -> None:
    """Deterministic output contract: stdout/stderr are always UTF-8.

    On Windows, a piped subprocess stdout defaults to the ANSI codepage
    (charmap) and fails on CJK text. The gateway reads JSON from stdout, so
    the encoding is part of the script contract.
    """
    for stream in (sys.stdout, sys.stderr):
        if hasattr(stream, "reconfigure"):
            stream.reconfigure(encoding="utf-8", errors="replace")


def main() -> None:
    _force_utf8_stdio()
    params: dict = {}
    if not sys.stdin.isatty():
        raw = sys.stdin.read()
        if raw.strip():
            try:
                params = json.loads(raw)
            except json.JSONDecodeError:
                params = {}
    if not params:
        parser = argparse.ArgumentParser(description="Workbook IR → styled XLSX dashboard")
        parser.add_argument("--input", required=True)
        parser.add_argument("--out", dest="output_dir", default="output")
        params = vars(parser.parse_args())
    try:
        run(params)
    except Exception as exc:  # noqa: BLE001
        print(json.dumps({"success": False, "message": str(exc), "context": {}}, ensure_ascii=False))
        sys.exit(1)


if __name__ == "__main__":
    main()
