import importlib.util
import unittest
from io import BytesIO
from pathlib import Path

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[2]
SCRIPT = ROOT / "skills" / "office-generate-production-dashboard" / "scripts" / "generate_dashboard.py"


def load_generator():
    spec = importlib.util.spec_from_file_location("generate_dashboard", SCRIPT)
    if spec is None or spec.loader is None:
        raise RuntimeError(f"unable to load {SCRIPT}")
    module = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(module)
    return module


class DashboardGenerationTests(unittest.TestCase):
    @classmethod
    def setUpClass(cls):
        cls.generator = load_generator()

    def test_a1_range_uses_the_start_cell_column(self):
        self.assertEqual(1, self.generator._ref_col("A2:A4"))
        self.assertEqual(2, self.generator._ref_col("B2:B4"))
        self.assertEqual((2, 4), self.generator._ref_rows("B2:B4"))

    def test_showcase_workbook_has_real_chart_refs_and_print_contract(self):
        ir = self.generator.load_workbook_ir(ROOT / "showcase" / "production-dashboard" / "input.json")
        visual_asset = ROOT / "showcase" / "production-dashboard" / "assets" / "data-landscape.jpg"
        workbook = self.generator.DashboardBuilder(ir, visual_asset=visual_asset).build()
        stream = BytesIO()
        workbook.save(stream)
        stream.seek(0)
        loaded = load_workbook(stream, data_only=False)
        runtime = loaded["Runtime Surface"]
        workflow = loaded["Workflow Packs"]

        self.assertEqual("A10", runtime.freeze_panes)
        self.assertEqual("landscape", runtime.page_setup.orientation)
        self.assertEqual(1, len(runtime._charts))
        chart = runtime._charts[0]
        self.assertEqual("'Runtime Surface'!$B$10:$B$12", chart.series[0].val.numRef.f)
        self.assertEqual("'Runtime Surface'!$A$10:$A$12", chart.series[0].cat.numRef.f)
        self.assertEqual("3157F6", chart.series[0].graphicalProperties.solidFill.srgbClr)
        self.assertTrue(chart.dLbls.showCatName)
        self.assertFalse(chart.dLbls.showSerName)
        self.assertEqual("=SUM(B10:B12)", runtime["G5"].value)
        self.assertEqual(1, len(runtime._images))
        self.assertEqual("Aptos Display", runtime["A1"].font.name)
        self.assertEqual("01\nSOURCE\nReusable IR", workflow["F10"].value)
        self.assertIn("Runtime_Surface_runtime_surface", runtime.tables)
        self.assertIn("Workflow_Packs_workflow_packs", workflow.tables)


if __name__ == "__main__":
    unittest.main()
